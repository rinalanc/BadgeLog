import streamlit as st
import os
import openpyxl
from openpyxl.utils import get_column_letter
import io # Para lidar com arquivos em memória
import zipfile # Para compactar as imagens para download

# --- IMPORTAÇÕES DAS BIBLIOTERAS DE CÓDIGOS DE BARRAS ---
from aztec_code_generator import AztecCode
from PIL import Image 
# --- FIM DAS IMPORTAÇÕES ---

def generate_aztec_codes_streamlit(input_text_content):
    """
    Gera códigos Aztec e um arquivo Excel a partir de um conteúdo de texto,
    retornando os dados em memória.

    Args:
        input_text_content (str): Conteúdo de texto de entrada (códigos separados por linha).

    Returns:
        tuple: (excel_buffer, image_buffers)
               excel_buffer (io.BytesIO): Buffer contendo o arquivo Excel.
               image_buffers (list): Lista de tuplas (filename, io.BytesIO) para as imagens.
    """
    # Divide o conteúdo da caixa de texto em linhas
    lines = input_text_content.splitlines()

    excel_data = []
    excel_data.append(["Dado Original", "Tipo de Código", "Nome do Arquivo Gerado"])

    image_buffers = [] # Lista para armazenar as imagens em memória

    # Filtra linhas vazias que podem surgir de quebras de linha extras
    lines = [line.strip() for line in lines if line.strip()]

    if not lines:
        st.warning("A caixa de texto está vazia ou contém apenas espaços. Nenhum código Aztec para gerar.")
        return None, []

    st.info(f"Gerando códigos para {len(lines)} entradas...")

    for i, line_data in enumerate(lines):
        code_data = line_data.strip()

        generated_filename = None
        image_buffer = io.BytesIO() # Buffer para a imagem atual
        
        try:
            aztec_code = AztecCode(code_data)
            generated_filename = f"aztec_code_{i+1}.png"
            # Salva no buffer em vez de um arquivo no disco, especificando o formato PNG
            aztec_code.save(image_buffer, format="PNG", module_size=4, border=1)
            st.write(f"Aztec Code gerado para '{code_data}'")

            if generated_filename:
                # Resetar o ponteiro do buffer para o início antes de ler ou adicionar ao ZIP
                image_buffer.seek(0)
                image_buffers.append((generated_filename, image_buffer))
                excel_data.append([code_data, "AZTEC", generated_filename])

        except Exception as e:
            st.error(f"Erro ao gerar Aztec Code para '{code_data}' (linha {i+1}): {e}")
            # Em Streamlit, queremos ver esses erros na interface

    # Cria o arquivo Excel em memória
    excel_buffer = io.BytesIO()
    if len(excel_data) > 1: # Verifica se há dados além do cabeçalho
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Códigos Aztec Gerados" 

            for row_data in excel_data:
                sheet.append(row_data)

            for col in range(1, sheet.max_column + 1):
                sheet.column_dimensions[get_column_letter(col)].auto_size = True

            workbook.save(excel_buffer)
            excel_buffer.seek(0) # Resetar o ponteiro para o início
            st.success("Arquivo Excel criado com sucesso!")
        except Exception as e:
            st.error(f"Erro ao criar o arquivo Excel: {e}")
            excel_buffer = None # Indica que o Excel não foi criado
    else:
        st.warning("Nenhum dado válido para ser salvo no arquivo Excel.")
        excel_buffer = None

    return excel_buffer, image_buffers


# --- Interface do Streamlit ---
st.set_page_config(
    page_title="Prometheus Aztec Generator",
    page_icon="logo.png", # Define o ícone da aba do navegador
    layout="centered"
)

# --- ALTERAÇÃO AQUI: Usando st.columns e st.image para o logo e st.markdown para o título ---
col1, col2 = st.columns([0.1, 0.9]) # Divide a largura em duas colunas (10% para imagem, 90% para texto)

with col1:
    st.image("logo.png", width=50) # Exibe a imagem do logo

with col2:
    st.markdown(
        f"<span style='font-size: 2.5em; font-weight: bold; vertical-align: middle;'>Prometheus Aztec Generator</span>",
        unsafe_allow_html=True
    )
# --- FIM DA ALTERAÇÃO ---

st.markdown("---")

st.markdown("""
Esta ferramenta gera **Códigos Aztec** a partir de dados inseridos diretamente.
""")

# 1. Caixa de texto para entrada de dados
st.header("1. Inserir Dados")
input_data_text = st.text_area(
    "Digite os códigos a serem gerados (um código por linha):",
    height=200,
    placeholder="Exemplo:\nCODIGO123\nPRODUTOABC\n4567890"
)

# Botão para iniciar a geração
st.header("2. Gerar Códigos Aztec e Excel")
if st.button("Gerar Códigos Aztec"):
    if input_data_text.strip(): # Verifica se a caixa de texto não está vazia ou só com espaços
        with st.spinner("Gerando códigos Aztec e arquivo Excel..."):
            excel_buffer, image_buffers = generate_aztec_codes_streamlit(input_data_text)
        
        st.markdown("---")
        st.header("3. Download dos Resultados")

        # Botão de download do Excel
        if excel_buffer:
            st.download_button(
                label="📥 Baixar Arquivo Excel (Aztec)",
                data=excel_buffer,
                file_name=f"aztec_codes_summary.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        # Botão de download das Imagens (compactadas em ZIP)
        if image_buffers:
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                for filename, img_buf in image_buffers:
                    zf.writestr(filename, img_buf.getvalue())
            
            zip_buffer.seek(0) # Resetar o ponteiro
            st.download_button(
                label="📦 Baixar Imagens Aztec (ZIP)",
                data=zip_buffer,
                file_name=f"aztec_images.zip",
                mime="application/zip"
            )
        
        if not excel_buffer and not image_buffers:
            st.warning("Nenhum arquivo gerado. Verifique a entrada e as mensagens de erro acima.")

    else:
        st.warning("Por favor, digite os códigos na caixa de texto para começar.")

st.markdown("---")
st.info("Desenvolvido com o apoio de Gemini (Google AI). Data: 02/07/2025")
